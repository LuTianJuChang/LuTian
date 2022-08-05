from flask_script import Command
import openpyxl
import os
from app.config import BASE_DIR
from app import db
from app.Models import Course, Sale


class ImportData(Command):
    def run(self):
        print('导入开始')
        dir = os.path.join(BASE_DIR, 'excel')
        for file_name in os.listdir(dir):
            file_path = os.path.join(dir, file_name)
            self.save_to_mysql(file_path)
        print('导入完成')

    def save_to_mysql(self, file_path):
        fields = ['course_id', 'product_id', 'product_type', 'product_name', 'provider', 'score', 'score_level',
                  'learner_count', 'lesson_count', 'lector_name', 'original_price', 'discount_price', 'discount_rate', 'img_url', 'big_img_url', 'description']

        # 导入excel文件
        wb = openpyxl.load_workbook(file_path)
        # 选择worksheet
        index = wb.sheetnames
        ws = wb[index[0]]
        for row in ws.iter_rows(min_row=2):
            # 从min_row最小行数到最大行数，从min_col最小列到最大列
            data = [cell.value for cell in row]
            dict_val = dict(zip(fields, data))
            # 存入数据表
            course = Course(**dict_val)
            self.save_course(course)
            create_time = file_path.split('.')[0][-16:]
            sale = Sale(course_id=dict_val['course_id'], product_name=dict_val['product_name'],
                        learner_count=dict_val['learner_count'], create_time=create_time)
            self.save_sale(sale, create_time)

            # 关闭excel连接
            wb.close()

    def save_course(self, course):
        try:
            db.session.merge(course)
            db.session.commit()
        except:
            db.session.rollback()

    def save_sale(self, sale, create_time):
        data = Sale.query.filter_by(course_id=sale.course_id, create_time=create_time).first()
        try:
            if not data:
                db.session.merge(sale)
                db.session.commit()
        except:
            db.session.rollback()

